
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
import os
import os
from google import genai
from pathlib import Path
from PIL import Image
from django.shortcuts import render, get_object_or_404
from django.views.decorators.http import require_POST
from .models import Message
from google import genai
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.contrib.auth import authenticate, login, logout
from .models import Room, Topic, Message, User, Announcement, EventInterest
from .forms import RoomForm, UserForm, MyUserCreationForm
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from .models import EventInterest
import pytesseract
from .supabase_storage import upload_file
import fitz 
import requests
import time
from io import BytesIO

# Create your views here.

# rooms = [
#     {'id': 1, 'name': 'Lets learn python!'},
#     {'id': 2, 'name': 'Design with me'},
#     {'id': 3, 'name': 'Frontend developers'},
# ]
def test_template(request):
    # Get the latest message (or create one for testing)
    message = Message.objects.last()  # or Message.objects.create(body="Test")
    
    context = {"message": message}
    return render(request, "base/test.html", context)

def dashboard_view(request):
    return render(request, 'base/index.html')

#announcemnt


# def ann_home(request):
#     return render(request, 'base/announcement.html' )

# views.py
# def event_interest(request, ann_id):
#     if request.method == 'POST':
#         roll_no = request.POST['roll_no']
#         course = request.POST['course']
#         mobile = request.POST['mobile']
#         email = request.POST['email']
#         # save interest or send notification etc.
#         messages.success(request, "Your interest has been recorded.")
#         return redirect('announcements')

# def announcements(request):
#     announcements = Announcement.objects.all().order_by('-created_at')
#     return render(request, 'base/announcement.html', {'announcements': announcements})
# --------------------------------------------------------------------------------------------------------------------

# @login_required(login_url='login')
# def announcements(request):

#     announcements = Announcement.objects.all().order_by('-created_at')
    
#     # Apply filters
#     search_query = request.GET.get('search')
#     school_filter = request.GET.get('school')
#     start_date = request.GET.get('start_date')
#     end_date = request.GET.get('end_date')
    
#     if search_query:
#         announcements = announcements.filter(title__icontains=search_query)
    
#     if school_filter:
#         announcements = announcements.filter(school_name=school_filter)
    
#     if start_date:
#         announcements = announcements.filter(event_date__gte=start_date)
    
#     if end_date:
#         announcements = announcements.filter(event_date__lte=end_date)




#     if request.method == "POST":
#         # Create new announcement
#         title = request.POST.get("title")
#         venue = request.POST.get("venue")
#         event_date = request.POST.get("event_date")
#         event_time = request.POST.get("event_time")
#         school_name = request.POST.get("school_name")
#         content = request.POST.get("content")

#         Announcement.objects.create(
#             author=request.user,
#             title=title,
#             venue=venue,
#             event_date=event_date,
#             event_time=event_time,
#             school_name=school_name,
#             content=content
#         )

#         messages.success(request, "Announcement posted successfully!")
#         return redirect("announcements")

#     # Existing: fetch all announcements
#     announcements = Announcement.objects.all().order_by('-created_at')
#     return render(request, 'base/announcement.html', {'announcements': announcements})
# -------------------------------------------------------------------------------------------------
@login_required(login_url='login')
def announcements(request):
    announcements = Announcement.objects.all().order_by('-created_at')
    
    # Apply filters
    search_query = request.GET.get('search')
    school_filter = request.GET.get('school')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if search_query:
        announcements = announcements.filter(title__icontains=search_query)
    
    if school_filter:
        announcements = announcements.filter(school_name=school_filter)
    
    if start_date:
        announcements = announcements.filter(event_date__gte=start_date)
    
    if end_date:
        announcements = announcements.filter(event_date__lte=end_date)

    # Get unique school names for dropdown
    schools = Announcement.objects.values_list('school_name', flat=True).distinct()

    if request.method == "POST":
        # Your existing POST handling code...
        title = request.POST.get("title")
        venue = request.POST.get("venue")
        event_date = request.POST.get("event_date")
        event_time = request.POST.get("event_time")
        school_name = request.POST.get("school_name")
        content = request.POST.get("content")

        Announcement.objects.create(
            author=request.user,
            title=title,
            venue=venue,
            event_date=event_date,
            event_time=event_time,
            school_name=school_name,
            content=content
        )
        messages.success(request, "Announcement posted successfully!")
        return redirect("announcements")

    return render(request, 'base/announcement.html', {
        'announcements': announcements,
        'schools': schools
    })

def event_interest(request, ann_id):
    if request.method == 'POST':
        announcement = get_object_or_404(Announcement, id=ann_id)
        roll_no = request.POST['roll_no']
        course = request.POST['course']
        mobile = request.POST['mobile']
        email = request.POST['email']

        EventInterest.objects.create(
            announcement=announcement,
            roll_no=roll_no,
            course=course,
            mobile=mobile,
            email=email
        )

        messages.success(request, "Your interest has been recorded successfully!")
        return redirect('announcements')

    return redirect('announcements')


#delete announcemnt
@login_required(login_url='login')
def delete_announcement(request, id):
    ann = get_object_or_404(Announcement, id=id)

    # Only author can delete
    if ann.author != request.user:
        messages.error(request, "You are not allowed to delete this announcement.")
        return redirect('announcements')

    ann.delete()
    messages.success(request, "Announcement deleted successfully!")
    return redirect('announcements')

#update announcement
@login_required(login_url='login')
def edit_announcement(request, id):
    ann = get_object_or_404(Announcement, id=id)

    if ann.author != request.user:
        messages.error(request, "You are not allowed to edit this announcement.")
        return redirect('announcements')

    if request.method == "POST":
        ann.title = request.POST.get("title")
        ann.venue = request.POST.get("venue")
        ann.event_date = request.POST.get("event_date")
        ann.event_time = request.POST.get("event_time")
        ann.school_name = request.POST.get("school_name")
        ann.content = request.POST.get("content")
        ann.save()

        messages.success(request, "Announcement updated successfully!")
        return redirect('announcements')

    return redirect('announcements')


def announcementProfile(request, pk):
    user = User.objects.get(id=pk)
    
    announcements = Announcement.objects.filter(author=user).order_by('-created_at')
    context = {
        'user': user,
        'announcements': announcements
    }
    return render(request, 'base/announcement_profile.html', context)





# def export_event_interests(request):
#     # Create a workbook and sheet
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Interested Students"

#     # Define headers
#     headers = ["S.No", "Announcement Title", "Roll No", "Course", "Mobile", "Email"]

#     # Style header
#     header_font = Font(bold=True, color="FFFFFF")
#     header_fill = PatternFill("solid", fgColor="4F81BD")
#     center_align = Alignment(horizontal="center", vertical="center")

#     # Write headers
#     for col_num, column_title in enumerate(headers, 1):
#         cell = ws.cell(row=1, column=col_num, value=column_title)
#         cell.font = header_font
#         cell.fill = header_fill
#         cell.alignment = center_align

#     # Fetch data
#     interests = EventInterest.objects.select_related('announcement').all().order_by('-id')

#     for row_num, interest in enumerate(interests, start=2):
#         ws.cell(row=row_num, column=1, value=row_num - 1)
#         ws.cell(row=row_num, column=2, value=interest.announcement.title)
#         ws.cell(row=row_num, column=3, value=interest.roll_no)
#         ws.cell(row=row_num, column=4, value=interest.course)
#         ws.cell(row=row_num, column=5, value=interest.mobile)
#         ws.cell(row=row_num, column=6, value=interest.email)
        

#     # Adjust column widths
#     for column_cells in ws.columns:
#         length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
#         ws.column_dimensions[column_cells[0].column_letter].width = length + 3

#     # Prepare response
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="interested_students.xlsx"'

#     wb.save(response)
#     return response

def export_event_interests(request, ann_id):
    announcement = get_object_or_404(Announcement, id=ann_id)

    interests = EventInterest.objects.filter(announcement=announcement).order_by('-id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Interested Students"

    headers = ["S.No", "Roll No", "Course", "Mobile", "Email"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    interests = EventInterest.objects.filter(announcement=announcement)

    for i, item in enumerate(interests, start=2):
        ws.cell(i, 1, i - 1)
        ws.cell(i, 2, item.roll_no)
        ws.cell(i, 3, item.course)
        ws.cell(i, 4, item.mobile)
        ws.cell(i, 5, item.email)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{announcement.title}_interests.xlsx"'
    wb.save(response)
    return response



def loginPage(request):
    page = 'login'
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        email = request.POST.get('email').lower()
        password = request.POST.get('password')

        try:
            user = User.objects.get(email=email)
        except:
            messages.error(request, 'User does not exist')

        user = authenticate(request, email=email, password=password)

        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, 'Username OR password does not exit')

    context = {'page': page}
    return render(request, 'base/login_register.html', context)


def logoutUser(request):
    logout(request)
    return redirect('home')



# def registerPage(request):
#     form = MyUserCreationForm()

#     if request.method == 'POST':
#         form = MyUserCreationForm(request.POST)
#         if form.is_valid():
#             user = form.save(commit=False)
#             user.username = user.username.lower()
#             user.save()
#             login(request, user)
#             return redirect('home')
#         else:
#             messages.error(request, 'An error occurred during registration')

#     return render(request, 'base/login_register.html', {'form': form})

def registerPage(request):
    form = MyUserCreationForm()

    if request.method == 'POST':
        form = MyUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.username = user.username.lower()
            user.save()
            login(request, user)
            return redirect('home')
        else:
            # Loop through all field errors and show them
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.capitalize()}: {error}")

    return render(request, 'base/login_register.html', {'form': form})


@login_required(login_url='login')
def home(request):
    q = request.GET.get('q') if request.GET.get('q') != None else ''

    rooms = Room.objects.filter(
        Q(topic__name__icontains=q) |
        Q(name__icontains=q) |
        Q(description__icontains=q)
    )

    topics = Topic.objects.all()[0:5]
    room_count = rooms.count()
    room_messages = Message.objects.filter(
        Q(room__topic__name__icontains=q))[0:3]

    context = {'rooms': rooms, 'topics': topics,
               'room_count': room_count, 'room_messages': room_messages}
    return render(request, 'base/home.html', context)


# def room(request, pk):
#     room = Room.objects.get(id=pk)
#     room_messages = room.message_set.all()
#     participants = room.participants.all()
#     hashtag_filter = request.GET.get('hashtag')

#     if request.method == 'POST':
#         message = Message.objects.create(
#             user=request.user,
#             room=room,
#             body=request.POST.get('body')
#         )
#         room.participants.add(request.user)
#         return redirect('room', pk=room.id)

#     context = {'room': room, 'room_messages': room_messages,
#                'participants': participants}
#     return render(request, 'base/room.html', context)

# def room(request, pk):
#     room = Room.objects.get(id=pk)
#     room_messages = room.message_set.all()
#     participants = room.participants.all()
#     hashtag_filter = request.GET.get('hashtag')

#     if hashtag_filter:
#         room_messages = room.message_set.filter(hashtags__icontains=hashtag_filter)
#     else:
#         room_messages = room.message_set.all()

#     participants = room.participants.all()

#     if request.method == 'POST':
#         message = Message.objects.create(
#             user=request.user,
#             room=room,
#             body=request.POST.get('body')
#         )
#         room.participants.add(request.user)
#         return redirect('room', pk=room.id)

#     # extract all hashtags in the room (unique)
#     all_tags = set()
#     for msg in room.message_set.all():
#         all_tags.update(msg.hashtag_list())

#     context = {
#         'room': room,
#         'room_messages': room_messages,
#         'participants': participants,
#         'hashtags': sorted(all_tags),
#     }
#     return render(request, 'base/room.html', context)
@login_required(login_url='login')
def room(request, pk):
    room = Room.objects.get(id=pk)
    room_messages = room.message_set.all()
    participants = room.participants.all()

    hashtag_filter1 = request.GET.get('hashtag', '').strip()
    hashtag_filter = hashtag_filter1.lstrip('#')

    if hashtag_filter:
        room_messages = room_messages.filter(hashtags__regex=rf'(^| )({hashtag_filter})( |$)')

    # Collect all distinct hashtags in the room
    all_hashtags = []
    for msg in room.message_set.all():
        all_hashtags.extend(msg.hashtags.split())
    distinct_hashtags = sorted(set(all_hashtags))

    if request.method == 'POST':
        body = request.POST.get('body')
        image = request.FILES.get('image')
        pdf = request.FILES.get('pdf')

        # Create message with files - let the model's save() handle Supabase upload
        if body or image or pdf:
            message = Message.objects.create(
                user=request.user,
                room=room,
                body=body,
                image=image,  # Pass the file object directly
                pdf=pdf      # Pass the file object directly
            )
            room.participants.add(request.user)

        return redirect('room', pk=room.id)

    context = {
        'room': room,
        'room_messages': room_messages,
        'participants': participants,
        'hashtags': distinct_hashtags,
        'hashtag_filter': hashtag_filter
    }
    return render(request, 'base/room.html', context)
# --------------------------------------------
# def room(request, pk):
#     room = Room.objects.get(id=pk)
#     room_messages = room.message_set.all()
#     participants = room.participants.all()

#     # hashtag_filter = request.GET.get('hashtag', '')
#     hashtag_filter1 = request.GET.get('hashtag', '').strip()
#     hashtag_filter = hashtag_filter1.lstrip('#')
#     # if not hashtag_filter.startswith('#') and hashtag_filter != '':
#     #     hashtag_filter = '#' + hashtag_filter


#     if hashtag_filter:
#         # room_messages = room_messages.filter(hashtags__icontains=hashtag_filter)
#         # room_messages = room_messages.filter(hashtags__regex=rf'\b{hashtag_filter}\b')
#         room_messages = room_messages.filter(hashtags__regex=rf'(^| )({hashtag_filter})( |$)')
        


#     # Collect all distinct hashtags in the room
#     all_hashtags = []
#     for msg in room.message_set.all():
#         all_hashtags.extend(msg.hashtags.split())
#     distinct_hashtags = sorted(set(all_hashtags))

#     # if request.method == 'POST':
#     #     message = Message.objects.create(
#     #         user=request.user,
#     #         room=room,
#     #         body=request.POST.get('body')                      # working one
#     #     )
#     #     room.participants.add(request.user)
#     #     return redirect('room', pk=room.id)
#     # body = None
#     # image = None
#     # pdf = None
    
#     # if request.method == 'POST':
#     #     body = request.POST.get('body')
#     #     image = request.FILES.get('image')
#     #     pdf = request.FILES.get('pdf')

#     #     if body or image or pdf:  # prevent empty messages
#     #         message=Message.objects.create(
#     #             user=request.user,
#     #             room=room,
#     #             body=body,
#     #             image=image,
#     #             pdf=pdf
#     #         )
#     #         room.participants.add(request.user)
            
#     #     return redirect('room', pk=room.id)
   

#     if request.method == 'POST':
#         body = request.POST.get('body')
#         image = request.FILES.get('image')
#         pdf = request.FILES.get('pdf')

#         image_url = None
#         pdf_url = None

#         # Upload image to supabase
#         if image:
#             image_url = upload_file(image, f"message_images/{image.name}")

#         # Upload pdf to supabase
#         if pdf:
#             pdf_url = upload_file(pdf, f"message_pdfs/{pdf.name}")

#         # Save message in database
#         if body or image_url or pdf_url:
#             message = Message.objects.create(
#                 user=request.user,
#                 room=room,
#                 body=body,
#                 image=image_url,   # store Supabase URL
#                 pdf=pdf_url        # store Supabase URL
#             )
#             room.participants.add(request.user)

#         return redirect('room', pk=room.id)



#     context = {
#         'room': room,
#         'room_messages': room_messages,
#         'participants': participants,
#         'hashtags': distinct_hashtags,
#         'hashtag_filter': hashtag_filter
#     }
#     return render(request, 'base/room.html', context)
# ////-------------------------------------------------------------



def userProfile(request, pk):
    user = User.objects.get(id=pk)
    rooms = user.room_set.all()
    room_messages = user.message_set.all()
    topics = Topic.objects.all()
    context = {'user': user, 'rooms': rooms,
               'room_messages': room_messages, 'topics': topics}
    return render(request, 'base/profile.html', context)


@login_required(login_url='login')
def createRoom(request):
    form = RoomForm()
    topics = Topic.objects.all()
    if request.method == 'POST':
        topic_name = request.POST.get('topic')
        topic, created = Topic.objects.get_or_create(name=topic_name)

        Room.objects.create(
            host=request.user,
            topic=topic,
            name=request.POST.get('name'),
            description=request.POST.get('description'),
        )
        return redirect('home')

    context = {'form': form, 'topics': topics}
    return render(request, 'base/room_form.html', context)


@login_required(login_url='login')
def updateRoom(request, pk):
    room = Room.objects.get(id=pk)
    form = RoomForm(instance=room)
    topics = Topic.objects.all()
    if request.user != room.host:
        return HttpResponse('Your are not allowed here!!')

    if request.method == 'POST':
        topic_name = request.POST.get('topic')
        topic, created = Topic.objects.get_or_create(name=topic_name)
        room.name = request.POST.get('name')
        room.topic = topic
        room.description = request.POST.get('description')
        room.save()
        return redirect('home')

    context = {'form': form, 'topics': topics, 'room': room}
    return render(request, 'base/room_form.html', context)


@login_required(login_url='login')
def deleteRoom(request, pk):
    room = Room.objects.get(id=pk)

    if request.user != room.host:
        return HttpResponse('Your are not allowed here!!')

    if request.method == 'POST':
        room.delete()
        return redirect('home')
    return render(request, 'base/delete.html', {'obj': room})


@login_required(login_url='login')
def deleteMessage(request, pk):
    message = Message.objects.get(id=pk)

    if request.user != message.user:
        return HttpResponse('Your are not allowed here!!')

    if request.method == 'POST':
        message.delete()
        return redirect('home')
    return render(request, 'base/delete.html', {'obj': message})


@login_required(login_url='login')
def updateUser(request):
    user = request.user
    form = UserForm(instance=user)

    if request.method == 'POST':
        form = UserForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            return redirect('user-profile', pk=user.id)
    

    return render(request, 'base/update-user.html', {'form': form})

@login_required(login_url='login')
def topicsPage(request):
    q = request.GET.get('q') if request.GET.get('q') != None else ''
    topics = Topic.objects.filter(name__icontains=q)
    return render(request, 'base/topics.html', {'topics': topics})

@login_required(login_url='login')
def activityPage(request):
    room_messages = Message.objects.all()
    return render(request, 'base/activity.html', {'room_messages': room_messages})



# Setup Gemini API client

os.environ["API_KEY"] = "AIzaSyDuCmFicMCoCmg2DUuT5eUZhIEPz_2f03c"
API_KEY = os.getenv("API_KEY")

client = genai.Client(api_key=os.environ["API_KEY"])

# def generate_flashcard(request, message_id):
#     message = get_object_or_404(Message, id=message_id)
#     input_text = message.body

#     # prompt = (
#     #     f"Read the following text and generate 3 flashcards in Q&A format. "
#     #     f"Each card should include a question and an answer covering key ideas.\n\n"
#     #     f"Text:\n{input_text}\n\n"
#     #     f"Flashcards:\n"
#     #     f"1. Q: ...\n   A: ...\n"
#     # )
#     prompt = (
#     f"Read the following text and generate 3 short, summary-style flashcards. "
#     f"Each flashcard should be concise so it's easy to read and understand quickly. "
#     f"Format each flashcard as a question and answer (Q&A), and add '###' at the end of each flashcard as a separator.\n\n"
#     f"Text:\n{input_text}\n\n"
#     f"Flashcards:\n"
#     f"1. Q: ...\n   A: ... ###\n"
# )


#     response = client.models.generate_content(
#         model="gemini-2.5-flash",
#         contents=prompt
#     )

#     generated_text = response.text if hasattr(response, "text") else "No flashcards generated."

#  # Process flashcards into a list of dicts: [{'Q': '...', 'A': '...'}, ...]
#     flashcards = []
#     for card_text in generated_text.split('\n\n'):
#         card_text = card_text.strip()
#         if not card_text:
#             continue
#         q, a = '', ''
#         for line in card_text.splitlines():
#             if line.startswith('Q:'):
#                 q = line[3:].strip()
#             elif line.startswith('A:'):
#                 a = line[3:].strip()
#         if q and a:
#             flashcards.append({'Q': q, 'A': a})

#     return render(request, "base/flashcards.html", {
#         "message": message,
#         "flashcards": flashcards
#     })

#=----------------------------------------------------------------------
# views.py

# Set your Gemini API key
# os.environ["API_KEY"] = "YOUR_GEMINI_API_KEY"
# client = genai.Client(api_key=os.environ["API_KEY"])

# Optional: Extract text from PDFs using PyMuPDF
# def extract_text_from_pdf(pdf_path):
#     import fitz  # PyMuPDF
#     doc = fitz.open(pdf_path)
#     text = ""                                                                        #this 1
#     for page in doc:
#         text += page.get_text()
#     return text




def extract_text_from_pdf_url(pdf_url):
    response = requests.get(pdf_url)
    doc = fitz.open(stream=response.content, filetype="pdf")
    
    text = ""
    for page in doc:
        text += page.get_text()

    return text


# Optional: Extract text from image
# def extract_text_from_image(image_path):
#     try:
#         image = Image.open(image_path)
#         text = pytesseract.image_to_string(image)
#         return text                                                                   #2
#     except Exception as e:
#         print("OCR failed:", e)
#         return ""


def extract_text_from_image_url(image_url):
    try:
        response = requests.get(image_url)
        image = Image.open(BytesIO(response.content))
        text = pytesseract.image_to_string(image)
        return text
    except Exception as e:
        print("OCR failed:", e)
        return ""




def generate_flashcards_from_text(input_text, num_cards=10, explain=True):
    prompt = (
        f"Read the following text carefully and generate **student-friendly flashcards**. "
        f"Requirements:\n"
        f"1. Each flashcard must cover **only one main idea**.\n"
        f"2. Keep flashcards **short and easy to remember** (1–2 sentences max).\n"
        f"3. Use **concise statements or bullets**; do NOT use Q&A format.\n"
        f"4. Each flashcard should be **self-contained**—it should make sense on its own.\n"
        f"5. Cover **all key points** from the text; create as many flashcards as necessary.\n"
        f"6. After each flashcard, include '###' as a separator.\n"
        f"7. Do not include any extra instructions or filler text.\n\n"
        f"Text:\n{input_text}\n\n"
        f"Flashcards:\n"
    )

    # Retry logic (up to 3 times)
    max_retries = 3
    for attempt in range(1, max_retries + 1):
        try:
            response = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=prompt
            )
            return response.text
        except Exception as e:
            if attempt < max_retries:
                print(f"⚠️ Attempt {attempt} failed: {e}. Retrying...")
                time.sleep(2)  # small delay before retry
            else:
                print(f"❌ All attempts failed: {e}")
                return None  # indicate failure


# def generate_flashcard(request, message_id):
#     message = get_object_or_404(Message, id=message_id)
#     input_text = ""

#     # 1. Text
#     if message.body:
#         input_text += message.body + "\n\n"

#     # 2. PDF
#     if message.pdf:
#         pdf_path = message.pdf.path
#         if Path(pdf_path).exists():
#             input_text += extract_text_from_pdf(pdf_path) + "\n\n"

#     # 3. Image
#     if message.image:
#         image_path = message.image.path
#         if Path(image_path).exists():
#             input_text += extract_text_from_image(image_path) + "\n\n"

#     flashcards = []
#     error_message = None

#     if input_text.strip():
#         flashcard_text = generate_flashcards_from_text(input_text)
#         if flashcard_text:
#             flashcards = [f.strip() for f in flashcard_text.split("###") if f.strip()]
#         else:
#             error_message = (
#                 "The flashcard service is temporarily overloaded. Please try again later."
#             )
#     else:
#         error_message = "No valid text found to generate flashcards."

#     context = {
#         "message": message,
#         "flashcards": flashcards,
#         "error_message": error_message,
#     }
#     return render(request, "base/flashcards.html", context)


def generate_flashcard(request, message_id):
    message = get_object_or_404(Message, id=message_id)
    input_text = ""

    # 1. Text
    if message.body:
        input_text += message.body + "\n\n"

    # 2. PDF (Supabase Storage URL)
    if message.pdf_url:
        pdf_url = message.pdf_url  # PUBLIC SUPABASE URL
        input_text += extract_text_from_pdf_url(pdf_url) + "\n\n"

    # 3. Image (Supabase Storage URL)
    if message.image_url:
        image_url = message.image_url  # PUBLIC SUPABASE URL
        input_text += extract_text_from_image_url(image_url) + "\n\n"

    flashcards = []
    error_message = None

    if input_text.strip():
        flashcard_text = generate_flashcards_from_text(input_text)
        if flashcard_text:
            flashcards = [f.strip() for f in flashcard_text.split("###") if f.strip()]
        else:
            error_message = (
                "The flashcard service is temporarily overloaded. Please try again later."
            )
    else:
        error_message = "No valid text found to generate flashcards."

    return render(request, "base/flashcards.html", {
        "message": message,
        "flashcards": flashcards,
        "error_message": error_message,
    })



#--------------------------------------------------------------------------------
# Generate flashcards using Gemini API
# def generate_flashcards_from_text(input_text, num_cards=10, explain=True):
#     # explanation = "Explain each answer clearly and concisely." if explain else ""
#     prompt = (
#     f"Read the following text carefully and generate **student-friendly flashcards**. "
#     f"Requirements:\n"
#     f"1. Each flashcard must cover **only one main idea**.\n"
#     f"2. Keep flashcards **short and easy to remember** (1–2 sentences max).\n"
#     f"3. Use **concise statements or bullets**; do NOT use Q&A format.\n"
#     f"4. Each flashcard should be **self-contained**—it should make sense on its own.\n"
#     f"5. Cover **all key points** from the text; create as many flashcards as necessary.\n"
#     f"6. After each flashcard, include '###' as a separator.\n"
#     f"7. Do not include any extra instructions or filler text.\n\n"
#     f"Text:\n{input_text}\n\n"
#     f"Flashcards:\n"
# )

#     response = client.models.generate_content(
#         model="gemini-2.5-flash",
#         contents=prompt
#     )
#     return response.text

# # Django view
# def generate_flashcard(request, message_id):
#     message = get_object_or_404(Message, id=message_id)
#     input_text = ""

#     # 1. Text input
#     if message.body:
#         input_text += message.body + "\n\n"

#     # 2. PDF input
#     if message.pdf:
#         pdf_path = message.pdf.path
#         if Path(pdf_path).exists():
#             input_text += extract_text_from_pdf(pdf_path) + "\n\n"

#     # 3. Image input
#     if message.image:
#         image_path = message.image.path
#         if Path(image_path).exists():
#             input_text += extract_text_from_image(image_path) + "\n\n"

#     flashcards = ""
#     if input_text.strip():  # Only generate if there's any text
#         flashcard_text = generate_flashcards_from_text(input_text, num_cards=10)
#         flashcards = [f.strip() for f in flashcard_text.split("###") if f.strip()]


#     context = {
#         "message": message,
#         "flashcards": flashcards,
#     }
#     return render(request, "base/flashcards.html", context)

#-------------------------------------------------------------------------------------11/11/25
# def extract_text_from_image(image_field):
#     """Extract text from an uploaded image using pytesseract"""
#     if not image_field:
#         return ""
#     try:
#         img = Image.open(image_field)
#         text = pytesseract.image_to_string(img)
#         return text.strip()
#     except Exception as e:
#         print("Image OCR error:", e)
#         return ""

# def extract_text_from_pdf(pdf_field):
#     """Extract text from a PDF (supports text PDFs and scanned PDFs)"""
#     if not pdf_field:
#         return ""
#     try:
#         pdf_path = pdf_field.path
#         doc = fitz.open(pdf_path)
#         full_text = ""

#         for page in doc:
#             page_text = page.get_text().strip()
#             if page_text:  # Normal text PDF
#                 full_text += page_text + "\n"
#             else:  # Scanned PDF → OCR
#                 pix = page.get_pixmap()
#                 img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
#                 ocr_text = pytesseract.image_to_string(img)
#                 full_text += ocr_text + "\n"

#         return full_text.strip()
#     except Exception as e:
#         print("PDF extraction error:", e)
#         return ""


# @login_required(login_url='login')
# def generate_flashcard(request, message_id):
#     message = get_object_or_404(Message, id=message_id)

#     # 1️⃣ Collect all text sources
#     texts = []

#     if message.body and message.body.strip():
#         texts.append(message.body.strip())

#     if message.image:
#         img_text = extract_text_from_image(message.image)
#         if img_text:
#             texts.append(img_text)

#     if message.pdf:
#         pdf_text = extract_text_from_pdf(message.pdf)
#         if pdf_text:
#             texts.append(pdf_text)

#     input_text = "\n".join(texts)

#     # 2️⃣ Handle empty content
#     if not input_text.strip():
#         flashcards = ["No readable text found in this message."]
#     else:
#         # 3️⃣ Build prompt for Gemini
#         prompt = (
#             f"Read the following text and generate flashcards in a student-friendly study style. "
#             f"Requirements:\n"
#             f"1. Each flashcard should cover **one main idea** only.\n"
#             f"2. Flashcards must be **short, 1–2 sentences max**.\n"
#             f"3. Do **NOT** use Q&A format. Just concise statements or bullets.\n"
#             f"4. Each flashcard should be **self-contained**.\n"
#             f"5. Include **all key points** from the text; create as many flashcards as needed.\n"
#             f"6. After each flashcard, add '###' as a separator.\n\n"
#             f"Text:\n{input_text}\n\n"
#             f"Flashcards:\n"
#         )

#         # 4️⃣ Call Gemini API
#         response = client.models.generate_content(
#             model="gemini-2.5-flash",
#             contents=prompt
#         )

#         flashcard_text = getattr(response, "text", "No flashcards generated.")
#         flashcards = [f.strip() for f in flashcard_text.split("###") if f.strip()]

#     # 5️⃣ Send to template
#     context = {
#         "message": message,
#         "flashcards": flashcards,
#     }
#     return render(request, "base/flashcards.html", context)
# -----------------------------------------------------------
# def extract_text_from_image(image_field):
#     """Extract text from an uploaded image"""
#     if not image_field:
#         return ""
#     try:
#         img = Image.open(image_field)
#         return pytesseract.image_to_string(img)
#     except Exception as e:
#         print("Image OCR error:", e)
#         return ""


# def extract_text_from_pdf(pdf_field):
#     """Extract text from an uploaded PDF"""
#     if not pdf_field:
#         return ""
#     try:
#         pdf_path = pdf_field.path  # path to the uploaded file
#         doc = fitz.open(pdf_path)
#         text = ""
#         for page in doc:
#             text += page.get_text()
#         return text
#     except Exception as e:
#         print("PDF extraction error:", e)
#         return ""


# def generate_flashcard(request, message_id):
#     message = get_object_or_404(Message, id=message_id)

#     # 1️⃣ Gather all text sources
#     input_text = message.body or ""
    
#     # Add OCR text if image exists
#     if message.image:
#         input_text += "\n" + extract_text_from_image(message.image)

#     # Add PDF text if PDF exists
#     if message.pdf:
#         input_text += "\n" + extract_text_from_pdf(message.pdf)

#     if not input_text.strip():
#         flashcards = ["No text available in this message."]
#     else:
#         # 2️⃣ Build Gemini prompt
#         prompt = (
#             f"Read the following text and generate flashcards in a student-friendly study style. "
#             f"Requirements:\n"
#             f"1. Each flashcard should cover **one main idea** only.\n"
#             f"2. Flashcards must be **short, 1–2 sentences max**.\n"
#             f"3. Do **NOT** use Q&A format. Just concise statements or bullets.\n"
#             f"4. Each flashcard should be **self-contained**.\n"
#             f"5. Include **all key points** from the text; create as many flashcards as needed.\n"
#             f"6. After each flashcard, add '###' as a separator.\n\n"
#             f"Text:\n{input_text}\n\n"
#             f"Flashcards:\n"
#         )

#         # 3️⃣ Call Gemini API
#         response = client.models.generate_content(
#             model="gemini-2.5-flash",
#             contents=prompt
#         )

#         # 4️⃣ Extract and split flashcards
#         flashcard_text = getattr(response, "text", "No flashcards generated.")
#         flashcards = [f.strip() for f in flashcard_text.split("###") if f.strip()]

#     # 5️⃣ Send to template
#     context = {
#         "message": message,
#         "flashcards": flashcards,
#     }
#     return render(request, "base/flashcards.html", context)

#-------------------------------------------------------------------------------------
# def generate_flashcard(request, message_id):
#     message = get_object_or_404(Message, id=message_id)

#     # Start with text from message body
#     input_text = message.body or ""

#     # --- Extract text from PDF if present ---
#     if message.pdf:
#         pdf_path = message.pdf.path
#         try:
#             # Convert PDF pages to images
#             pages = convert_from_path(pdf_path)
#             for page in pages:
#                 input_text += "\n" + pytesseract.image_to_string(page)
#         except Exception as e:
#             input_text += f"\n[Error reading PDF: {str(e)}]"

#     # --- Extract text from image if present ---
#     if message.image:
#         image_path = message.image.path
#         try:
#             img = Image.open(image_path)
#             input_text += "\n" + pytesseract.image_to_string(img)
#         except Exception as e:
#             input_text += f"\n[Error reading image: {str(e)}]"

#     # --- Build prompt for flashcard generation ---
#     prompt = (
#         f"Read the following text and generate flashcards in a student-friendly study style. "
#         f"Requirements:\n"
#         f"1. Each flashcard should cover **one main idea** only.\n"
#         f"2. Flashcards must be **short, 1–2 sentences max**, so they are easy to read and remember.\n"
#         f"3. Do **NOT** use Q&A format. Just concise statements or bullets.\n"
#         f"4. Each flashcard should be **self-contained**—it should make sense on its own.\n"
#         f"5. Include **all key points** from the text; create as many flashcards as needed.\n"
#         f"6. After each flashcard, add '###' as a separator.\n\n"
#         f"Text:\n{input_text}\n\n"
#         f"Flashcards:\n"
#     )

#     # --- Call Gemini API ---
#     response = client.models.generate_content(
#         model="gemini-2.5-flash",
#         contents=prompt
#     )

#     # Extract generated text
#     flashcard_text = getattr(response, "text", "No flashcards generated.")

#     # Split flashcards by '###' separator
#     flashcards = [f.strip() for f in flashcard_text.split("###") if f.strip()]

#     # Send to template using context
#     context = {
#         'message': message,
#         'flashcards': flashcards
#     }

#     return render(request, 'base/flashcards.html', context)


#old
# def generate_flashcard(request, message_id):
#     message = get_object_or_404(Message, id=message_id)
#     input_text = message.body

#     # Your prompt for short, summary-style flashcards with separator
#     prompt = (
#         f"Read the following text and generate flashcards in a student-friendly study style. "
#         f"Requirements:\n"
#         f"1. Each flashcard should cover **one main idea** only.\n"
#         f"2. Flashcards must be **short, 1–2 sentences max**, so they are easy to read and remember.\n"
#         f"3. Do **NOT** use Q&A format. Just concise statements or bullets.\n"
#         f"4. Each flashcard should be **self-contained**—it should make sense on its own.\n"
#         f"5. Include **all key points** from the text; create as many flashcards as needed.\n"
#         f"6. After each flashcard, add '###' as a separator.\n\n"
#         f"Text:\n{input_text}\n\n"
#         f"Flashcards:\n"
#     )


#     # prompt = (
#     #     f"Read the following text and generate short, summary-style flashcards for all important points. "
#     #     f"Each flashcard should be concise, highlighting a single main idea so that anyone can quickly understand the content. "
#     #     f"Do NOT limit the number of flashcards—create as many as needed to cover all key points. "
#     #     f"Do NOT use Q&A format, just short clear statements. "
#     #     f"Add '###' at the end of each flashcard as a separator.\n\n"
#     #     f"Text:\n{input_text}\n\n"
#     #     f"Flashcards:\n"
#     # )


#     # Call to Gemini API
#     response = client.models.generate_content(
#         model="gemini-2.5-flash",
#         contents=prompt
#     )

#     # Extract generated text
#     flashcard_text = response.text if hasattr(response, "text") else "No flashcards generated."
#   # or check your API response structure

#     # Split flashcards by '###' separator
#     flashcards = [f.strip() for f in flashcard_text.split("###") if f.strip()]

#     # Send to template using context
#     context = {
#         'message': message,
#         'flashcards': flashcards
#     }

#     return render(request, 'base/flashcards.html', context)


# old 
# in views.py


# @login_required(login_url='login')
# @require_POST
# def add_hashtag(request, message_id):
#     message = get_object_or_404(Message, id=message_id)
#     if request.user != message.user:
#         return HttpResponse('You are not allowed to tag this message.')

#     new_tag = request.POST.get('hashtag', '').strip()
#     if not new_tag.startswith('#'):
#         new_tag = f'#{new_tag}'

#     # Append to existing hashtags
#     existing_tags = message.hashtag_list()
#     if new_tag not in existing_tags:
#         existing_tags.append(new_tag)
#         message.hashtags = " ".join(existing_tags)
#         message.save()

#     return redirect('room', pk=message.room.id)

@login_required(login_url='login')
@require_POST
def add_hashtag(request, message_id):
    message = get_object_or_404(Message, id=message_id)
    if request.user != message.user:
        return HttpResponse('You are not allowed to tag this message.')

    # Get user input and clean it
    new_tag = request.POST.get('hashtag', '').strip()

    # ✅ Remove # if user typed it by mistake
    if new_tag.startswith('#'):
        new_tag = new_tag[1:]

    # Append to existing hashtags
    existing_tags = message.hashtag_list()
    if new_tag and new_tag not in existing_tags:
        existing_tags.append(new_tag)
        message.hashtags = " ".join(existing_tags)
        message.save()

    return redirect('room', pk=message.room.id)



# @login_required(login_url='login')
# def edit_message_hashtags(request, pk):
#     message = get_object_or_404(Message, id=pk)

#     # Only the message owner can edit
#     if request.user != message.user:
#         return HttpResponse("You are not allowed to edit this message!")

#     if request.method == "POST":
#         new_hashtags = request.POST.get("hashtags", "")
#         message.hashtags = new_hashtags
#         message.save()
#         return redirect('room', pk=message.room.id)

#     return render(request, "base\edit_hashtags.html", {"message": message})
@login_required(login_url='login')
def edit_message_hashtags(request, pk):
    message = get_object_or_404(Message, id=pk)

    # Only the message owner can edit
    if request.user != message.user:
        return HttpResponse("You are not allowed to edit this message!")

    if request.method == "POST":
        new_hashtags = request.POST.get("hashtags", "")
        message.hashtags = new_hashtags
        message.save()
        return redirect('room', pk=message.room.id)

    return render(request, "base/edit_hashtags.html", {"message": message})




